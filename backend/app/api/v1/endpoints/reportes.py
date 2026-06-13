"""
Endpoints para el módulo de Reportes
"""
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, List
from datetime import date
from io import BytesIO

from app.core.deps import get_db, get_current_user
from app.models.usuario import Usuario
from app.services import reportes_service
from app.schemas.reportes import (
    ReporteVentasMaterial,
    ReporteVentasSemanal,
    ReporteGastos,
    ReporteMaquinaria,
    ReporteTopClientes,
    ResumenGeneral,
    ReporteMovimientos
)

router = APIRouter()


@router.get("/materiales", response_model=List[str])
def obtener_materiales(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Obtiene la lista de materiales disponibles para filtrar"""
    return reportes_service.obtener_materiales_disponibles(db)


@router.get("/resumen", response_model=ResumenGeneral)
def obtener_resumen_general(
    fecha_desde: Optional[date] = Query(None, description="Fecha inicio del período"),
    fecha_hasta: Optional[date] = Query(None, description="Fecha fin del período"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Obtiene un resumen general del período:
    - Total de ventas (kg, toneladas, importe)
    - Total de gastos por categoría
    - Máquinas activas
    - Cliente y material top
    """
    return reportes_service.obtener_resumen_general(db, fecha_desde, fecha_hasta)


@router.get("/ventas-material", response_model=ReporteVentasMaterial)
def obtener_ventas_por_material(
    fecha_desde: Optional[date] = Query(None, description="Fecha inicio del período"),
    fecha_hasta: Optional[date] = Query(None, description="Fecha fin del período"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Obtiene las ventas agrupadas por tipo de material/árido:
    - Cantidad de pesajes por material
    - Total de kg y toneladas
    - Importe total
    - Precio promedio
    """
    return reportes_service.obtener_ventas_por_material(db, fecha_desde, fecha_hasta)


@router.get("/ventas-semanal", response_model=ReporteVentasSemanal)
def obtener_ventas_semanales(
    fecha_desde: Optional[date] = Query(None, description="Fecha inicio del período"),
    fecha_hasta: Optional[date] = Query(None, description="Fecha fin del período"),
    material: Optional[str] = Query(None, description="Filtrar por material específico"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Obtiene las ventas agrupadas por semana:
    - Total vendido por semana
    - Filtrable por tipo de material
    """
    return reportes_service.obtener_ventas_semanales(db, fecha_desde, fecha_hasta, material)


@router.get("/gastos", response_model=ReporteGastos)
def obtener_reporte_gastos(
    fecha_desde: Optional[date] = Query(None, description="Fecha inicio del período"),
    fecha_hasta: Optional[date] = Query(None, description="Fecha fin del período"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Obtiene el reporte de gastos:
    - Combustible
    - Servicios/Mantenimiento
    - Repuestos
    - Otros gastos
    """
    return reportes_service.obtener_reporte_gastos(db, fecha_desde, fecha_hasta)


@router.get("/maquinaria", response_model=ReporteMaquinaria)
def obtener_reporte_maquinaria(
    fecha_desde: Optional[date] = Query(None, description="Fecha inicio del período"),
    fecha_hasta: Optional[date] = Query(None, description="Fecha fin del período"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Obtiene el reporte de actividad de maquinaria:
    - Cantidad de máquinas que trabajaron
    - Pesajes por máquina
    - Combustible consumido
    - Servicios realizados
    """
    return reportes_service.obtener_reporte_maquinaria(db, fecha_desde, fecha_hasta)


@router.get("/top-clientes", response_model=ReporteTopClientes)
def obtener_top_clientes(
    fecha_desde: Optional[date] = Query(None, description="Fecha inicio del período"),
    fecha_hasta: Optional[date] = Query(None, description="Fecha fin del período"),
    limite: int = Query(10, description="Cantidad de clientes a mostrar"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Obtiene los clientes con más compras:
    - Ordenados por total de kg
    - Con importe y porcentaje del total
    """
    return reportes_service.obtener_top_clientes(db, fecha_desde, fecha_hasta, limite)


@router.get("/movimientos", response_model=ReporteMovimientos)
def obtener_movimientos(
    fecha_desde: Optional[date] = Query(None, description="Fecha inicio del período"),
    fecha_hasta: Optional[date] = Query(None, description="Fecha fin del período"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Obtiene el detalle de todos los pesajes en el período.
    Ideal para cierre de caja y control de movimientos diarios.
    """
    return reportes_service.obtener_movimientos(db, fecha_desde, fecha_hasta)


@router.get("/movimientos/exportar-excel")
def exportar_movimientos_excel(
    fecha_desde: Optional[date] = Query(None, description="Fecha inicio del período"),
    fecha_hasta: Optional[date] = Query(None, description="Fecha fin del período"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Exporta los movimientos/pesajes a Excel
    """
    import csv
    from io import StringIO

    reporte = reportes_service.obtener_movimientos(db, fecha_desde, fecha_hasta)

    # Crear CSV en memoria
    output = StringIO()
    writer = csv.writer(output)

    # Encabezados
    writer.writerow([
        'N° Pesaje', 'Fecha', 'Hora', 'Cliente', 'Material',
        'Peso Neto (kg)', 'Precio/Tn', 'Flete', 'Importe Total',
        'Patente', 'Chofer', 'Tipo Entrega', 'Estado', 'Observaciones'
    ])

    # Datos
    for m in reporte.movimientos:
        writer.writerow([
            m.numero_pesaje,
            m.fecha,
            m.hora,
            m.cliente_nombre or '',
            m.material or '',
            float(m.peso_neto) if m.peso_neto else '',
            float(m.precio_unitario) if m.precio_unitario else '',
            float(m.flete) if m.flete else '',
            float(m.importe_total) if m.importe_total else '',
            m.patente or '',
            m.chofer or '',
            m.tipo_entrega or '',
            m.estado,
            m.observaciones or ''
        ])

    # Fila de totales
    writer.writerow([])
    writer.writerow(['TOTALES', '', '', '', '',
                     float(reporte.total_kg), '', '',
                     float(reporte.total_importe), '', '', '', '', ''])
    writer.writerow([f'Cantidad de pesajes: {reporte.cantidad_pesajes}'])
    writer.writerow([f'Total toneladas: {float(reporte.total_toneladas):.2f}'])

    # Preparar respuesta
    output.seek(0)
    content = output.getvalue().encode('utf-8-sig')  # BOM para Excel

    fecha_str = f"{reporte.fecha_desde.strftime('%Y%m%d')}_{reporte.fecha_hasta.strftime('%Y%m%d')}"
    filename = f"movimientos_{fecha_str}.csv"

    return StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/movimientos/exportar-xlsx")
def exportar_movimientos_xlsx(
    fecha_desde: Optional[date] = Query(None, description="Fecha inicio del período"),
    fecha_hasta: Optional[date] = Query(None, description="Fecha fin del período"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Exporta los movimientos/pesajes a Excel real (.xlsx) con openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    reporte = reportes_service.obtener_movimientos(db, fecha_desde, fecha_hasta)

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    headers = [
        "N° Pesaje", "Fecha", "Hora", "Cliente", "Material",
        "Peso Neto (kg)", "Precio/Tn", "Flete", "Importe Total",
        "Patente", "Chofer", "Tipo Entrega", "Estado", "Observaciones",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for m in reporte.movimientos:
        ws.append([
            m.numero_pesaje,
            m.fecha,
            m.hora,
            m.cliente_nombre or "",
            m.material or "",
            float(m.peso_neto) if m.peso_neto else None,
            float(m.precio_unitario) if m.precio_unitario else None,
            float(m.flete) if m.flete else None,
            float(m.importe_total) if m.importe_total else None,
            m.patente or "",
            m.chofer or "",
            m.tipo_entrega or "",
            m.estado,
            m.observaciones or "",
        ])

    ws.append([])
    totales_row = [
        "TOTALES", "", "", "", "",
        float(reporte.total_kg) if reporte.total_kg else 0,
        "", "",
        float(reporte.total_importe) if reporte.total_importe else 0,
        "", "", "", "", "",
    ]
    ws.append(totales_row)
    for col in range(1, len(totales_row) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    ws.append([f"Cantidad de pesajes: {reporte.cantidad_pesajes}"])
    ws.append([f"Total toneladas: {float(reporte.total_toneladas):.2f}"])

    # Anchos automáticos básicos.
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            14, len(str(headers[col_idx - 1])) + 2
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fecha_str = f"{reporte.fecha_desde.strftime('%Y%m%d')}_{reporte.fecha_hasta.strftime('%Y%m%d')}"
    filename = f"movimientos_{fecha_str}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/exportar-pdf")
def exportar_reporte_pdf(
    tipo: str = Query(..., description="Tipo de reporte: resumen, ventas-material, ventas-semanal, gastos, maquinaria, top-clientes, movimientos"),
    fecha_desde: Optional[date] = Query(None, description="Fecha inicio del período"),
    fecha_hasta: Optional[date] = Query(None, description="Fecha fin del período"),
    material: Optional[str] = Query(None, description="Material para filtro (solo ventas-semanal)"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Exporta un reporte a PDF
    """
    from app.services.reportes_pdf_service import generar_pdf_reporte

    pdf_buffer = generar_pdf_reporte(
        db=db,
        tipo=tipo,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        material=material
    )

    # Nombre del archivo
    fecha_str = date.today().strftime("%Y%m%d")
    filename = f"reporte_{tipo}_{fecha_str}.pdf"

    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
